import streamlit as st
import pandas as pd
from io import BytesIO
import re
from typing import Dict, Set, Tuple, List, Any

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None  # fallback: pandas only

st.set_page_config(page_title="Daily Log vs Transaction Report", layout="wide")
st.title("Daily Log vs Transaction Report Comparator")

# ----------------------
# Configuration
# ----------------------
# Threshold (bytes) above which we use streaming parsing instead of loading whole file with pandas.
STREAMING_THRESHOLD = 10 * 1024 * 1024  # 10 MB

# Candidate header variants (normalized)
ACCESSION_CANDIDATES = [
    "accession", "accessionnumber", "accessionno", "acc", "accno", "accnumber",
    "inv", "invno", "invnumber", "invoice", "inv#",
]
CODE_CANDIDATES = [
    "testcode", "test_code", "code", "cpt", "procedurecode", "procedure", "testid", "test"
]
TESTNAME_CANDIDATES = [
    "testname", "test_name", "testnamefield", "testdesc", "testdescription", "description", "test"
]


# ----------------------
# Helpers
# ----------------------
def normalize_header(h: Any) -> str:
    if h is None:
        return ""
    s = str(h).strip().lower()
    s = re.sub(r"[^a-z0-9]", "", s)
    return s


def find_header_index(headers: List[str], candidates: List[str]) -> int:
    """
    Return index of first header that matches any candidate (normalized).
    Returns -1 if none found.
    """
    norm_map = {i: normalize_header(h) for i, h in enumerate(headers)}
    # exact match with candidates
    candidate_norms = {re.sub(r"[^a-z0-9]", "", c.lower()): c for c in candidates}
    for i, nh in norm_map.items():
        if nh in candidate_norms:
            return i
    # contains keyword fallback (e.g., header contains 'inv' or 'code')
    for i, nh in norm_map.items():
        for cand in candidates:
            cand_norm = re.sub(r"[^a-z0-9]", "", cand.lower())
            if cand_norm and cand_norm in nh:
                return i
    return -1


def to_str_safe(value):
    if value is None:
        return ""
    return str(value).strip()


# ----------------------
# Streaming parsers (openpyxl)
# ----------------------
def stream_parse_excel(file_bytes: bytes):
    """
    Returns:
      headers (list)
      rows_generator -> yields lists of cell values (strings preserved)
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl not available; cannot stream large Excel files. Install openpyxl.")
    bio = BytesIO(file_bytes)
    wb = load_workbook(filename=bio, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return [], iter([])
    headers = [to_str_safe(h) for h in header_row]
    def gen():
        for r in rows:
            yield [to_str_safe(c) for c in r]
    return headers, gen()


def parse_transaction_stream(file_bytes: bytes):
    """
    Build txn_map: inv -> set(codes)
    Build txn_testname_map: (inv, code) -> testname (first encountered)
    """
    headers, rows = stream_parse_excel(file_bytes)
    if not headers:
        return {}, {}
    inv_idx = find_header_index(headers, ACCESSION_CANDIDATES)
    code_idx = find_header_index(headers, CODE_CANDIDATES)
    name_idx = find_header_index(headers, TESTNAME_CANDIDATES)

    if inv_idx == -1 or code_idx == -1:
        raise ValueError(f"Transaction file: required columns not found. Available headers: {headers}")

    txn_map: Dict[str, Set[str]] = {}
    txn_testname_map: Dict[Tuple[str, str], str] = {}

    for r in rows:
        inv = r[inv_idx] if inv_idx < len(r) else ""
        code = r[code_idx] if code_idx < len(r) else ""
        name = r[name_idx] if (name_idx != -1 and name_idx < len(r)) else ""
        inv = to_str_safe(inv)
        code = to_str_safe(code)
        name = to_str_safe(name)
        if not inv:
            continue
        codeset = txn_map.setdefault(inv, set())
        if code:
            codeset.add(code)
            key = (inv, code)
            if key not in txn_testname_map and name:
                txn_testname_map[key] = name
    return txn_map, txn_testname_map, headers


def parse_daily_stream(file_bytes: bytes):
    """
    Build daily_map: accession -> set(codes)
    Build daily_template_map: accession -> first full-row dict (header->value)
    Returns (daily_map, daily_template_map, headers)
    """
    headers, rows = stream_parse_excel(file_bytes)
    if not headers:
        return {}, {}, headers

    acc_idx = find_header_index(headers, ACCESSION_CANDIDATES)
    code_idx = find_header_index(headers, CODE_CANDIDATES)

    if acc_idx == -1 or code_idx == -1:
        raise ValueError(f"Daily Log file: required columns not found. Available headers: {headers}")

    daily_map: Dict[str, Set[str]] = {}
    daily_template_map: Dict[str, Dict[str, str]] = {}

    for r in rows:
        # convert row into dict using headers
        row_values = [to_str_safe(c) for c in r]
        accession = row_values[acc_idx] if acc_idx < len(row_values) else ""
        code = row_values[code_idx] if code_idx < len(row_values) else ""
        accession = to_str_safe(accession)
        code = to_str_safe(code)
        if not accession:
            continue
        codeset = daily_map.setdefault(accession, set())
        if code:
            codeset.add(code)
        # save template row if first occurrence
        if accession not in daily_template_map:
            template = {}
            for i, h in enumerate(headers):
                template[h] = row_values[i] if i < len(row_values) else ""
            daily_template_map[accession] = template
    return daily_map, daily_template_map, headers


# ----------------------
# Pandas parsers (for smaller files)
# ----------------------
@st.cache_data(show_spinner=False)
def parse_transaction_pandas(file_bytes: bytes):
    bio = BytesIO(file_bytes)
    df = pd.read_excel(bio, dtype=str)
    # normalize column names
    cols = list(df.columns)
    acc_col = None
    code_col = None
    name_col = None
    for c in cols:
        nc = normalize_header(c)
        if acc_col is None and nc in [re.sub(r"[^a-z0-9]", "", x) for x in ACCESSION_CANDIDATES]:
            acc_col = c
        if code_col is None and nc in [re.sub(r"[^a-z0-9]", "", x) for x in CODE_CANDIDATES]:
            code_col = c
        if name_col is None and nc in [re.sub(r"[^a-z0-9]", "", x) for x in TESTNAME_CANDIDATES]:
            name_col = c
    # fallback heuristics
    if acc_col is None:
        for c in cols:
            if "inv" in normalize_header(c):
                acc_col = c; break
    if code_col is None:
        for c in cols:
            if "code" in normalize_header(c) or "cpt" in normalize_header(c) or "test" in normalize_header(c):
                code_col = c; break
    if acc_col is None or code_col is None:
        raise ValueError(f"Transaction file: required columns not found. Available headers: {cols}")

    df[acc_col] = df[acc_col].fillna("").astype(str).str.strip()
    df[code_col] = df[code_col].fillna("").astype(str).str.strip()

    txn_map = df.groupby(acc_col)[code_col].apply(lambda s: set(s[s != ""])).to_dict()
    txn_testname_map = {}
    if name_col and name_col in df.columns:
        # use first encountered name for each (inv, code)
        for _, row in df[[acc_col, code_col, name_col]].dropna(subset=[acc_col, code_col]).iterrows():
            key = (str(row[acc_col]).strip(), str(row[code_col]).strip())
            if key not in txn_testname_map:
                txn_testname_map[key] = str(row[name_col]).strip()
    return txn_map, txn_testname_map, list(df.columns)


@st.cache_data(show_spinner=False)
def parse_daily_pandas(file_bytes: bytes):
    bio = BytesIO(file_bytes)
    df = pd.read_excel(bio, dtype=str)
    cols = list(df.columns)
    acc_col = None
    code_col = None
    for c in cols:
        nc = normalize_header(c)
        if acc_col is None and nc in [re.sub(r"[^a-z0-9]", "", x) for x in ACCESSION_CANDIDATES]:
            acc_col = c
        if code_col is None and nc in [re.sub(r"[^a-z0-9]", "", x) for x in CODE_CANDIDATES]:
            code_col = c
    if acc_col is None:
        for c in cols:
            if "acc" in normalize_header(c) or "accession" in normalize_header(c) or "inv" in normalize_header(c):
                acc_col = c; break
    if code_col is None:
        for c in cols:
            if "code" in normalize_header(c) or "test" in normalize_header(c) or "cpt" in normalize_header(c):
                code_col = c; break
    if acc_col is None or code_col is None:
        raise ValueError(f"Daily Log file: required columns not found. Available headers: {cols}")

    df[acc_col] = df[acc_col].fillna("").astype(str).str.strip()
    df[code_col] = df[code_col].fillna("").astype(str).str.strip()

    daily_map = df.groupby(acc_col)[code_col].apply(lambda s: set(s[s != ""])).to_dict()

    # template map: first full-row dict per accession
    template_map = {}
    for acc, sub in df.groupby(acc_col):
        if acc not in template_map:
            # take first row as template
            first_row = sub.iloc[0].to_dict()
            # convert nan to ""
            template_map[str(acc)] = {k: ("" if pd.isna(v) else str(v)) for k, v in first_row.items()}
    return daily_map, template_map, list(df.columns), df


# ----------------------
# Main UI / Execution
# ----------------------
daily_file = st.file_uploader("Upload Daily Log Report", type=["xlsx", "xls"])
txn_file = st.file_uploader("Upload Transaction Report", type=["xlsx", "xls"])

if daily_file and txn_file:
    # read file bytes up-front (so we can inspect size and reuse bytes)
    daily_bytes = daily_file.read()
    txn_bytes = txn_file.read()

    daily_size = len(daily_bytes)
    txn_size = len(txn_bytes)

    use_stream_daily = daily_size > STREAMING_THRESHOLD
    use_stream_txn = txn_size > STREAMING_THRESHOLD

    st.subheader("Preview of uploaded files (headers only if large)")
    col1, col2 = st.columns(2)
    with col1:
        st.write(f"Daily Log — {daily_file.name} — {daily_size / 1024:.1f} KB")
    with col2:
        st.write(f"Transaction — {txn_file.name} — {txn_size / 1024:.1f} KB")

    # Safe parse with fallback and explicit error reporting
    try:
        if use_stream_txn:
            if load_workbook is None:
                raise RuntimeError("Transaction file is large and openpyxl is not available to stream it.")
            txn_map, txn_testname_map, txn_headers = parse_transaction_stream(txn_bytes)
            st.write("Transaction headers:", txn_headers)
        else:
            txn_map, txn_testname_map, txn_headers = parse_transaction_pandas(txn_bytes)
            # small preview
            # to avoid heavy memory use, only show first 5 rows of the DataFrame columns
            st.write("Transaction headers:", txn_headers)

        if use_stream_daily:
            if load_workbook is None:
                raise RuntimeError("Daily file is large and openpyxl is not available to stream it.")
            daily_map, daily_template_map, daily_headers = parse_daily_stream(daily_bytes)
            st.write("Daily Log headers:", daily_headers)
            # For template-row usage later we only have header names from stream; we will create DataFrame columns from daily_headers
            original_daily_df = None
        else:
            daily_map, daily_template_map, daily_headers, original_daily_df = parse_daily_pandas(daily_bytes)
            st.write("Daily Log headers (sample):", daily_headers)
            st.write("Daily Log preview", original_daily_df.head())

        # Build cleaned rows
        notes = []
        cleaned_rows = []

        # Process daily rows: keep only codes that exist in txn for same accession
        # If accession exists in txn:
        #   if code in txn_map[accession] -> keep (add Test Name if available)
        #   else -> mark removed (do not keep)
        # else:
        #   keep row and note accession missing in txn (kept)
        if original_daily_df is not None:
            # Faster vectorized path using pandas DataFrame (small daily file)
            acc_col = None
            code_col = None
            for c in original_daily_df.columns:
                nc = normalize_header(c)
                if acc_col is None and nc in [re.sub(r"[^a-z0-9]", "", x) for x in ACCESSION_CANDIDATES]:
                    acc_col = c
                if code_col is None and nc in [re.sub(r"[^a-z0-9]", "", x) for x in CODE_CANDIDATES]:
                    code_col = c
            if acc_col is None:
                for c in original_daily_df.columns:
                    if "acc" in normalize_header(c) or "inv" in normalize_header(c) or "accession" in normalize_header(c):
                        acc_col = c; break
            if code_col is None:
                for c in original_daily_df.columns:
                    if "code" in normalize_header(c) or "test" in normalize_header(c) or "cpt" in normalize_header(c):
                        code_col = c; break
            if acc_col is None or code_col is None:
                raise ValueError("Could not detect Accession/Test Code columns in Daily Log.")

            # iterate rows but keep original row structure
            for _, row in original_daily_df.iterrows():
                accession = to_str_safe(row[acc_col])
                code = to_str_safe(row[code_col])
                row_dict = {k: ("" if pd.isna(v) else str(v)) for k, v in row.to_dict().items()}

                if accession in txn_map:
                    if code and code in txn_map[accession]:
                        # keep
                        row_dict["Test Name"] = txn_testname_map.get((accession, code), row_dict.get("Test Name", ""))
                        cleaned_rows.append(row_dict)
                    else:
                        if code:
                            notes.append({"Accession": accession, "Action": f"Removed code {code}"})
                        else:
                            # if no code present, keep row
                            notes.append({"Accession": accession, "Action": "No code present in daily row (kept)"})
                            cleaned_rows.append(row_dict)
                else:
                    notes.append({"Accession": accession, "Action": "Accession not in Transaction Report (kept)"})
                    cleaned_rows.append(row_dict)
        else:
            # streaming daily: use daily_template_map and daily_map
            for accession, codes in daily_map.items():
                for code in codes:
                    row_dict = dict(daily_template_map.get(accession, {}))
                    if accession in txn_map:
                        if code in txn_map[accession]:
                            row_dict["Test Code"] = code
                            row_dict["Test Name"] = txn_testname_map.get((accession, code), row_dict.get("Test Name", ""))
                            cleaned_rows.append(row_dict)
                        else:
                            notes.append({"Accession": accession, "Action": f"Removed code {code}"})
                    else:
                        notes.append({"Accession": accession, "Action": "Accession not in Transaction Report (kept)"})
                        row_dict["Test Code"] = code
                        cleaned_rows.append(row_dict)

        # Add missing codes from Transaction Report
        # For each accession in txn_map, find txn_codes - daily_codes and add rows
        for accession, txn_codes in txn_map.items():
            daily_codes = daily_map.get(accession, set())
            missing_codes = txn_codes - daily_codes
            if missing_codes:
                for code in missing_codes:
                    # use template_row from daily_template_map if exists
                    if accession in daily_template_map:
                        template_row = dict(daily_template_map[accession])
                    else:
                        # if we have original_daily_df, try to build a minimal template with its columns
                        if original_daily_df is not None:
                            # take first row from entire daily df as template (to preserve all columns)
                            first_row = original_daily_df.iloc[0].to_dict()
                            template_row = {k: ("" if pd.isna(v) else str(v)) for k, v in first_row.items()}
                            template_row["Accession"] = accession
                        else:
                            # create minimal template using txn headers
                            template_row = {"Accession": accession}
                    template_row["Test Code"] = code
                    template_row["Test Name"] = txn_testname_map.get((accession, code), "")
                    cleaned_rows.append(template_row)
                    notes.append({"Accession": accession, "Action": f"Added missing code {code}"})

        # Final cleaned DataFrame
        cleaned_daily_df = pd.DataFrame(cleaned_rows)

        # TAB UI
        tab1, tab2, tab3 = st.tabs(["Cleaned Daily Log", "Discrepancy Notes", "Reconciliation"])

        with tab1:
            st.subheader("Cleaned Daily Log Report (preview)")
            st.dataframe(cleaned_daily_df.head(100))

            # Download option - create Excel in-memory efficiently
            try:
                output = BytesIO()
                # If DataFrame empty, create a default empty file with headers
                if cleaned_daily_df.empty:
                    pd.DataFrame(columns=["Accession", "Test Code", "Test Name"]).to_excel(output, index=False)
                else:
                    cleaned_daily_df.to_excel(output, index=False)
                st.download_button(
                    label="Download Cleaned Daily Log Report",
                    data=output.getvalue(),
                    file_name="Cleaned_Daily_Log_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Failed to prepare download: {e}")

        with tab2:
            st.subheader("Discrepancy Notes")
            notes_df = pd.DataFrame(notes)
            if not notes_df.empty:
                st.dataframe(notes_df)
            else:
                st.info("No discrepancies found")

        with tab3:
            st.subheader("Reconciliation Breakdown")

            # Accession level
            st.markdown("### Accession Level Reconciliation")
            daily_accessions = set(daily_map.keys())
            txn_accessions = set(txn_map.keys())

            missing_in_txn = daily_accessions - txn_accessions
            missing_in_daily = txn_accessions - daily_accessions

            accession_recon = pd.DataFrame({
                "Metric": ["Total in Daily Log", "Total in Transaction", "Missing in Transaction", "Missing in Daily Log"],
                "Count": [len(daily_accessions), len(txn_accessions), len(missing_in_txn), len(missing_in_daily)]
            })
            st.table(accession_recon)

            # CPT/Code Level per accession (show sample, limit rows)
            st.markdown("### CPT/Code Level Reconciliation (per Accession)")
            all_accessions = sorted(daily_accessions | txn_accessions)
            recon_rows = []
            # If too many accessions, limit to first 1000 to avoid UI overload
            MAX_RECON_ROWS = 1000
            for acc in all_accessions[:MAX_RECON_ROWS]:
                d_codes = daily_map.get(acc, set())
                t_codes = txn_map.get(acc, set())
                missing_in_txn_codes = d_codes - t_codes
                missing_in_daily_codes = t_codes - d_codes
                recon_rows.append({
                    "Accession": acc,
                    "Daily Codes": ", ".join(sorted(d_codes)) if d_codes else None,
                    "Txn Codes": ", ".join(sorted(t_codes)) if t_codes else None,
                    "Missing in Transaction": ", ".join(sorted(missing_in_txn_codes)) if missing_in_txn_codes else None,
                    "Missing in Daily Log": ", ".join(sorted(missing_in_daily_codes)) if missing_in_daily_codes else None
                })

            cpt_recon_df = pd.DataFrame(recon_rows)
            st.dataframe(cpt_recon_df)
    except Exception as exc:
        st.error("An error occurred while processing the files.")
        st.exception(exc)
